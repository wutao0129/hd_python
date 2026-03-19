"""
指标库管理 API (Issue #134-#137)
从 competency_model.py 拆分为独立路由
"""
import csv
import io
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import or_
from sqlalchemy.orm import Session
from database import get_db
from models import IndicatorLibrary, HrIndicatorLibraryLog
from middleware import get_current_user

router = APIRouter(prefix="/api/hr/indicator-library", tags=["指标库"])


# ---------- Schemas ----------

class IndicatorCreate(BaseModel):
    code: str = Field(..., max_length=50)
    name: str = Field(..., max_length=100)
    category: str = Field(..., max_length=50)
    weight: int = Field(default=0, ge=0, le=100)
    description: Optional[str] = None
    level: str = Field(default="熟练", max_length=20)
    is_preset: bool = False


class IndicatorUpdate(BaseModel):
    name: str = Field(..., max_length=100)
    category: str = Field(..., max_length=50)
    weight: int = Field(default=0, ge=0, le=100)
    description: Optional[str] = None
    level: Optional[str] = None


class StatusToggle(BaseModel):
    status: int = Field(..., ge=0, le=1)


# ---------- 辅助函数 ----------

def _to_dict(item: IndicatorLibrary) -> dict:
    return {
        "id": item.id,
        "code": item.code,
        "name": item.name,
        "description": item.description,
        "category": item.category,
        "weight": item.weight,
        "level": item.level,
        "isPreset": item.is_preset,
        "status": item.status,
        "refCount": item.ref_count,
        "createdBy": item.created_by,
        "updatedBy": item.updated_by,
        "createdAt": item.created_at.isoformat() if item.created_at else None,
        "updatedAt": item.updated_at.isoformat() if item.updated_at else None,
    }


def log_indicator_operation(
    db: Session,
    indicator_id: Optional[int],
    operation: str,
    detail: dict,
    operator: str,
):
    """记录操作日志"""
    log = HrIndicatorLibraryLog(
        indicator_id=indicator_id,
        operation=operation,
        detail=detail,
        operator=operator,
    )
    db.add(log)


# ---------- 查询接口 ----------

@router.get("")
def get_indicator_library(
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """分页查询指标库（支持多条件筛选）"""
    query = db.query(IndicatorLibrary).filter(IndicatorLibrary.deleted_at.is_(None))

    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(
            or_(
                IndicatorLibrary.code.like(kw),
                IndicatorLibrary.name.like(kw),
            )
        )
    if category:
        query = query.filter(IndicatorLibrary.category == category)
    if status is not None:
        query = query.filter(IndicatorLibrary.status == status)

    total = query.count()
    items = (
        query.order_by(IndicatorLibrary.id.desc())
        .offset((page - 1) * pageSize)
        .limit(pageSize)
        .all()
    )

    return {
        "code": 200,
        "message": "success",
        "data": {
            "total": total,
            "page": page,
            "pageSize": pageSize,
            "items": [_to_dict(item) for item in items],
        },
    }


@router.get("/logs")
def get_indicator_logs(
    page: int = Query(1, ge=1),
    pageSize: int = Query(20, ge=1, le=100),
    operation: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """查询操作日志"""
    query = db.query(HrIndicatorLibraryLog)

    if operation:
        query = query.filter(HrIndicatorLibraryLog.operation == operation)
    if start_date:
        query = query.filter(HrIndicatorLibraryLog.operated_at >= start_date)
    if end_date:
        query = query.filter(HrIndicatorLibraryLog.operated_at <= end_date)

    total = query.count()
    items = (
        query.order_by(HrIndicatorLibraryLog.operated_at.desc())
        .offset((page - 1) * pageSize)
        .limit(pageSize)
        .all()
    )

    return {
        "code": 200,
        "data": {
            "total": total,
            "page": page,
            "pageSize": pageSize,
            "items": [
                {
                    "id": log.id,
                    "indicatorId": log.indicator_id,
                    "operation": log.operation,
                    "detail": log.detail,
                    "operator": log.operator,
                    "operatedAt": log.operated_at.isoformat() if log.operated_at else None,
                }
                for log in items
            ],
        },
    }


@router.get("/check-code")
def check_code(
    code: str,
    exclude_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """检查指标编码唯一性"""
    query = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.code == code,
        IndicatorLibrary.deleted_at.is_(None),
    )
    if exclude_id:
        query = query.filter(IndicatorLibrary.id != exclude_id)
    exists = query.first() is not None
    return {"code": 200, "data": {"exists": exists}}


@router.get("/export")
def export_indicators(
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导出指标库为 Excel"""
    query = db.query(IndicatorLibrary).filter(IndicatorLibrary.deleted_at.is_(None))

    if keyword:
        kw = f"%{keyword}%"
        query = query.filter(or_(IndicatorLibrary.code.like(kw), IndicatorLibrary.name.like(kw)))
    if category:
        query = query.filter(IndicatorLibrary.category == category)
    if status is not None:
        query = query.filter(IndicatorLibrary.status == status)

    items = query.order_by(IndicatorLibrary.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "指标库"

    headers = ["指标编码", "指标名称", "所属维度", "推荐权重(%)", "引用次数", "状态", "指标描述"]
    ws.append(headers)

    for item in items:
        ws.append([
            item.code,
            item.name,
            item.category,
            item.weight,
            item.ref_count,
            "启用" if item.status == 1 else "禁用",
            item.description or "",
        ])

    for col, width in zip("ABCDEFG", [15, 20, 15, 15, 10, 10, 30]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.now().strftime("%Y%m%d")
    filename = f"indicator_library_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/import/template")
def download_import_template(
    current_user: dict = Depends(get_current_user),
):
    """下载指标导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "指标库导入模板"

    headers = ["指标编码", "指标名称", "所属维度", "推荐权重(%)", "指标描述"]
    ws.append(headers)

    ws.append(["IND-001", "市场分析能力", "业务能力", 30, "对市场趋势的分析能力"])
    ws.append(["IND-002", "沟通表达", "个人特质", 25, ""])

    for col, width in zip("ABCDE", [15, 20, 15, 15, 30]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=indicator_import_template.xlsx"},
    )


@router.get("/{indicator_id}")
def get_indicator_detail(
    indicator_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取单条指标详情"""
    item = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.id == indicator_id,
        IndicatorLibrary.deleted_at.is_(None),
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="指标不存在")
    return {"code": 200, "data": _to_dict(item)}


# ---------- 新增 ----------

@router.post("")
def create_indicator(
    data: IndicatorCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """新增指标"""
    # 编码唯一性校验
    existing = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.code == data.code,
        IndicatorLibrary.deleted_at.is_(None),
    ).first()
    if existing:
        return {"code": 400, "message": f"指标编码 '{data.code}' 已存在"}

    item = IndicatorLibrary(
        code=data.code,
        name=data.name,
        category=data.category,
        weight=data.weight,
        description=data.description,
        level=data.level,
        is_preset=data.is_preset,
        status=1,
        ref_count=0,
        created_by=current_user.get("username", ""),
        updated_by=current_user.get("username", ""),
    )
    db.add(item)
    db.flush()

    log_indicator_operation(db, item.id, "create", {"indicator": _to_dict(item)}, current_user.get("username", ""))
    db.commit()
    db.refresh(item)

    return {"code": 200, "message": "指标添加成功", "data": _to_dict(item)}


# ---------- 编辑 ----------

@router.put("/{indicator_id}")
def update_indicator(
    indicator_id: int,
    data: IndicatorUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """编辑指标（编码不可修改）"""
    item = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.id == indicator_id,
        IndicatorLibrary.deleted_at.is_(None),
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="指标不存在")

    before = {"name": item.name, "category": item.category, "weight": item.weight, "description": item.description}

    item.name = data.name
    item.category = data.category
    item.weight = data.weight
    item.description = data.description
    if data.level is not None:
        item.level = data.level
    item.updated_by = current_user.get("username", "")

    after = {"name": item.name, "category": item.category, "weight": item.weight, "description": item.description}
    log_indicator_operation(db, item.id, "update", {"before": before, "after": after}, current_user.get("username", ""))
    db.commit()
    db.refresh(item)

    return {"code": 200, "message": "指标更新成功", "data": _to_dict(item)}


# ---------- 删除 ----------

@router.delete("/{indicator_id}")
def delete_indicator(
    indicator_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """删除指标（软删除，引用保护）"""
    item = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.id == indicator_id,
        IndicatorLibrary.deleted_at.is_(None),
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="指标不存在")

    if item.ref_count > 0:
        return {"code": 400, "message": "该指标已被引用，无法删除"}

    item.deleted_at = datetime.utcnow()
    item.updated_by = current_user.get("username", "")

    log_indicator_operation(db, item.id, "delete", {"indicator": _to_dict(item)}, current_user.get("username", ""))
    db.commit()

    return {"code": 200, "message": "指标删除成功"}


# ---------- 启用/停用 ----------

@router.patch("/{indicator_id}/status")
def toggle_indicator_status(
    indicator_id: int,
    data: StatusToggle,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """启用/停用指标"""
    item = db.query(IndicatorLibrary).filter(
        IndicatorLibrary.id == indicator_id,
        IndicatorLibrary.deleted_at.is_(None),
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="指标不存在")

    # 停用时检查引用
    if data.status == 0 and item.ref_count > 0:
        return {"code": 400, "message": "该指标已被引用，无法停用"}

    old_status = item.status
    item.status = data.status
    item.updated_by = current_user.get("username", "")

    operation = "enable" if data.status == 1 else "disable"
    log_indicator_operation(
        db, item.id, operation,
        {"before": {"status": old_status}, "after": {"status": data.status}},
        current_user.get("username", ""),
    )
    db.commit()

    action = "启用" if data.status == 1 else "停用"
    return {"code": 200, "message": f"指标{action}成功", "data": _to_dict(item)}


# ---------- 导入辅助函数 ----------

VALID_CATEGORIES = {"业务能力", "专业技能", "领导力", "个人特质"}
INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize(value: str) -> str:
    """防止 CSV 注入"""
    if isinstance(value, str) and value and value[0] in INJECTION_PREFIXES:
        return "'" + value
    return value


def _parse_xlsx(file_bytes: bytes) -> List[dict]:
    """解析 Excel 文件"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    result = []
    for i, row in enumerate(rows):
        if not row or not any(row):
            continue
        code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        category = str(row[2] or "").strip()
        weight_raw = row[3]
        desc = str(row[4] or "").strip() if len(row) > 4 else ""
        result.append({
            "row": i + 2,
            "code": _sanitize(code),
            "name": _sanitize(name),
            "category": category,
            "weight": weight_raw,
            "description": _sanitize(desc),
        })
    wb.close()
    return result


def _parse_csv(file_bytes: bytes) -> List[dict]:
    """解析 CSV 文件"""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    result = []
    for i, row in enumerate(reader):
        result.append({
            "row": i + 2,
            "code": _sanitize(row.get("指标编码", "").strip()),
            "name": _sanitize(row.get("指标名称", "").strip()),
            "category": row.get("所属维度", "").strip(),
            "weight": row.get("推荐权重(%)", "0").strip(),
            "description": _sanitize(row.get("指标描述", "").strip()),
        })
    return result


# ---------- 批量导入 ----------

@router.post("/import")
async def import_indicators(
    file: UploadFile = File(...),
    strategy: str = Form("skip"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """批量导入指标"""
    if not file.filename:
        return {"code": 400, "message": "未选择文件"}

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "csv"):
        return {"code": 400, "message": "仅支持 .xlsx 和 .csv 文件"}

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        return {"code": 400, "message": "文件大小不能超过 10MB"}

    try:
        rows = _parse_xlsx(file_bytes) if ext == "xlsx" else _parse_csv(file_bytes)
    except Exception as e:
        return {"code": 400, "message": f"文件解析失败: {str(e)}"}

    if len(rows) > 5000:
        return {"code": 400, "message": "单次导入不超过 5000 条"}
    if not rows:
        return {"code": 400, "message": "文件中无有效数据"}

    success = 0
    skipped = 0
    failed = 0
    errors = []
    username = current_user.get("username", "")

    for row_data in rows:
        row_num = row_data["row"]
        code = row_data["code"]
        name = row_data["name"]
        category = row_data["category"]
        desc = row_data["description"]

        if not code:
            errors.append({"row": row_num, "reason": "指标编码不能为空"})
            failed += 1
            continue
        if not name:
            errors.append({"row": row_num, "reason": "指标名称不能为空"})
            failed += 1
            continue
        if category not in VALID_CATEGORIES:
            errors.append({"row": row_num, "reason": f"维度'{category}'无效"})
            failed += 1
            continue

        try:
            weight = int(row_data["weight"]) if row_data["weight"] else 0
        except (ValueError, TypeError):
            errors.append({"row": row_num, "reason": "权重必须为整数"})
            failed += 1
            continue

        if weight < 0 or weight > 100:
            errors.append({"row": row_num, "reason": "权重必须在 0-100 之间"})
            failed += 1
            continue

        existing = db.query(IndicatorLibrary).filter(
            IndicatorLibrary.code == code,
            IndicatorLibrary.deleted_at.is_(None),
        ).first()

        if existing:
            if strategy == "skip":
                skipped += 1
                continue
            else:
                existing.name = name
                existing.category = category
                existing.weight = weight
                existing.description = desc
                existing.updated_by = username
                success += 1
        else:
            db.add(IndicatorLibrary(
                code=code, name=name, category=category, weight=weight,
                description=desc, status=1, ref_count=0,
                created_by=username, updated_by=username,
            ))
            success += 1

    log_indicator_operation(
        db, None, "import",
        {"total": len(rows), "success": success, "skipped": skipped, "failed": failed, "strategy": strategy},
        username,
    )
    db.commit()

    return {
        "code": 200,
        "message": f"导入完成：成功 {success} 条，跳过 {skipped} 条，失败 {failed} 条",
        "data": {"success": success, "skipped": skipped, "failed": failed, "errors": errors[:50]},
    }
